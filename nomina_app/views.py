from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.template import loader
from .models import Dia, Semana, Venta, Tecnico, OtroGasto, PiezaAPagar
from .forms import DiaForm, SemanaForm, VentaForm, OtroGastoForm, PiezaAPagarForm
import datetime
#import xlwt
from django.contrib.auth.models import User
#from xlutils.copy import copy # http://pypi.python.org/pypi/xlutils
#from xlrd import open_workbook # http://pypi.python.org/pypi/xlrd

from openpyxl import load_workbook

import os

#TODO: adicionar ventas del dia al main
#todo: chequear que el dia no exista
# Create your views here.
def template(request):
    template = loader.get_template('nomina_app/base2.html')
    return HttpResponse(template.render({}, request))

def createSemana(request):
    template = loader.get_template('nomina_app/create-semana.html')
    context = getCurrentDate()
    now = datetime.datetime.now()
    semana=Semana()
    semana.nombre = now.strftime("%Y-%m-%d")
    form = SemanaForm(instance=semana)
    
    context['form']= form
    
    return HttpResponse(template.render(context, request))

def createNewSemana(request):
    # return HttpResponse("sdaasd")
    semana = Semana()
    semana.abierta = True
    new_semana = SemanaForm(request.POST, instance=semana)
    if new_semana.is_valid():
        new_semana.abierta = True
        new_semana.save()
    return redirect('main')

def createDia(request):
    semana = getCurrentDate()['semana']
    if not semana:
        return redirect('semana_abrir')
    template = loader.get_template('nomina_app/create-dia.html')
    context = getCurrentDate()
    now = datetime.datetime.now()
    dia=Dia()
    dia.fecha =datetime.datetime.now()
    # dia.nombre = now.strftime("%Y-%m-%d")
    form = DiaForm(instance=dia)
    
    context['form']= form
    
    return HttpResponse(template.render(context, request))

def createNewDia(request):
    # return HttpResponse("sdaasd")
    dia = Dia()
    dia.abierto = True
    dia.semana =getCurrentDate()['semana']
    
    new_dia = DiaForm(request.POST, instance=dia)
    if new_dia.is_valid():
        new_dia.abierta = True
        new_dia.save()
    return redirect('main')

def createNewVenta(request):
    # return HttpResponse("sdaasd")
    venta = Venta()
    venta.hora = datetime.datetime.now()
    venta.dia =getCurrentDate()['dia']
    
    new_venta = VentaForm(request.POST, instance=venta)
    if new_venta.is_valid():
        venta.costo = 0 if not venta.costo else venta.costo
        new_venta.save()
    return redirect('main')

def createNewGasto(request):
    # return HttpResponse("sdaasd")
    otroGasto = OtroGasto()
    otroGasto.dia =getCurrentDate()['dia']
    
    new_otroGasto = OtroGastoForm(request.POST, instance=otroGasto)
    new_otroGasto.save()
    return redirect('main')

def createNewPiezaAPagar(request):
    # return HttpResponse("sdaasd")
    piezaAPagar = PiezaAPagar()
    piezaAPagar.dia =getCurrentDate()['dia']
    
    new_piezaAPagar = PiezaAPagarForm(request.POST, instance=piezaAPagar)
    new_piezaAPagar.save()
    return redirect('main')

def colseDay(request):
    day= getCurrentDate()['dia']
    if day:
        day.abierto =False
        day.save()
    return redirect('dia_abrir')

def colseWeek(request):
    [semana, dia] = getCurrentDate()
    if not dia and semana:
        semana.abierta =False
        semana.save()
    return redirect('semana_abrir')

def exportToExcel(request):
    pass

def mainView(request):
    template = loader.get_template('nomina_app/main.html')
    context = getCurrentDate()
    dia = getCurrentDate()['dia']
    if not dia:
        return redirect('dia_abrir')
    #creating venta form
    venta = Venta()
    form = VentaForm(instance=venta)
    context['form']= form
    
    #creating OtroGasto Form
    otroGasto = OtroGasto()
    otroGastoForm = OtroGastoForm(instance=otroGasto)
    context['otroGastoForm']= otroGastoForm
    
    #creating PiezaAPagar Form
    piezaAPagar = PiezaAPagar()
    piezaAPagarForm = PiezaAPagarForm(instance=piezaAPagar)
    context['piezaAPagarForm']= piezaAPagarForm
    
    
    ventasList = 0
    if dia:
        ventasList = Venta.objects.filter(dia=dia)
    context['ventasList'] = ventasList
    
    otroGastoList = 0
    if dia:
        otroGastoList = OtroGasto.objects.filter(dia=dia)
    context['otroGastoList'] = otroGastoList
    
    piezasAPagarList= 0
    if dia:
        piezasAPagarList = PiezaAPagar.objects.filter(dia=dia)
    context['piezasAPagarList'] = piezasAPagarList
    
    return HttpResponse(template.render(context, request))

def exportToExcel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="users.xls"'

    path = os.path.dirname(__file__)
    file = os.path.join(path, 'excels/semana.xlsx')

    rb = load_workbook(file)
    
    wb = (rb)
    
    semana = getCurrentDate()['semana']
    
    if semana:
        days = Dia.objects.filter(semana=semana).order_by('id')
        sheetIndex = 0
        for day in days:
            #TODO: check index
            sheet = wb.worksheets[sheetIndex]
            sheetIndex+=1
            ventasList = Venta.objects.filter(dia=day)
            initialRow = 3
            for venta in ventasList:
                sheet[f'A{initialRow}']= venta.hora
                sheet[f'B{initialRow}']= venta.tipoTrabajo.__str__()
                sheet[f'C{initialRow}']= venta.nombre
                sheet[f'E{initialRow}']= venta.ingreso
                sheet[f'F{initialRow}']= venta.costo
                sheet[f'L{initialRow}']= venta.tecnico.__str__()
                initialRow+=1      
                
            otroGastoList = OtroGasto.objects.filter(dia=day)
            initialRow = 3
            cant =len(otroGastoList) //4
            i=0
            text = [[],[],[],[]]
            values= [[],[],[],[]]
            for gasto in otroGastoList:
                text[i%4].append(gasto.nombre);
                values[i%4].append(str(gasto.precio));                        
                i+=1
            sheet[f'C152']= ','.join(text[0])
            sheet[f'C153']= ','.join(text[1])
            sheet[f'C154']= ','.join(text[2])
            sheet[f'C155']= ','.join(text[3])
            
            sheet[f'D152']= '='+ '+'.join(values[0])
            sheet[f'D153']= '='+ '+'.join(values[1])
            sheet[f'D154']= '='+ '+'.join(values[2])
            sheet[f'D155']= '='+ '+'.join(values[3])
            
            piezaAPagar = PiezaAPagar.objects.filter(dia=day)
            initialRow = 3
            cant =len(piezaAPagar) //4
            i=0
            text = [[],[],[],[]]
            values= [[],[],[],[]]
            for pieza in piezaAPagar:
                text[i%4].append(pieza.nombre);
                values[i%4].append(str(pieza.precio));                        
                i+=1
            sheet[f'G152']= ','.join(text[0])
            sheet[f'E153']= ','.join(text[1])
            sheet[f'E154']= ','.join(text[2])
            sheet[f'E155']= ','.join(text[3])
            
            sheet[f'M152']= '='+ '+'.join(values[0])
            sheet[f'M153']= '='+ '+'.join(values[1])
            sheet[f'M154']= '='+ '+'.join(values[2])
            sheet[f'M155']= '='+ '+'.join(values[3])
            
        wb.save(response)
        return response
    return redirect('main')    
##################################################################################################
# wb = Workbook()
# dest_filename = 'empty_book.xlsx'
# ws1 = wb.active
# ws1.title = "range names"
# for row in range(1, 40):
#     ws1.append(range(600))
# ws2 = wb.create_sheet(title="Pi")
# ws2['F5'] = 3.14
# ws3 = wb.create_sheet(title="Data")
# for row in range(10, 20):
#     for col in range(27, 54):
#         _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
# print(ws3['AA10'].value)
# wb.save(filename = dest_filename)
# Read an existing workbook
##################################################################################################

def getCurrentDate():
    info ={}
    info['semana'] = Semana.objects.filter(abierta=True).first()
    info['dia']= Dia.objects.filter(abierto=True).first()
    return info
    